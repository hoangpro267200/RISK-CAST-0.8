"""
Excel export for AI-generated recommendations.

RISKCAST v17 - Export Recommendations to Excel

Provides professional Excel exports with:
- Risk summary section
- Formatted recommendations table
- Priority color coding
- Charts (optional)
"""

from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, Reference
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class RecommendationExporter:
    """
    Export AI recommendations to Excel format.
    
    Creates professional spreadsheets with:
    - Company branding
    - Risk assessment summary
    - Prioritized recommendations table
    - Action items checklist
    
    Usage:
        exporter = RecommendationExporter()
        excel_file = exporter.export_recommendations(
            recommendations=[...],
            risk_data={...}
        )
        
        # Return as FastAPI response
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=recommendations.xlsx"}
        )
    """
    
    # Color scheme
    COLORS = {
        'header': '366092',      # Dark blue
        'header_text': 'FFFFFF', # White
        'high': 'FF6B6B',        # Red
        'medium': 'FFD93D',      # Yellow
        'low': '6BCB77',         # Green
        'border': 'CCCCCC',      # Light gray
        'alt_row': 'F5F5F5'      # Very light gray
    }
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )
    
    def export_recommendations(
        self,
        recommendations: List[Dict[str, Any]],
        risk_data: Dict[str, Any],
        include_charts: bool = False,
        company_name: str = "RISKCAST"
    ) -> BytesIO:
        """
        Create Excel file with recommendations.
        
        Args:
            recommendations: List of recommendation dicts
            risk_data: Risk assessment data
            include_charts: Whether to include charts
            company_name: Company name for branding
        
        Returns:
            BytesIO containing Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "AI Recommendations"
        
        # Set column widths
        column_widths = {
            'A': 12,  # Priority
            'B': 15,  # Category
            'C': 50,  # Recommendation
            'D': 12,  # Impact
            'E': 12,  # Effort
            'F': 15,  # Cost Estimate
            'G': 15,  # Timeline
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        current_row = 1
        
        # === HEADER ===
        current_row = self._add_header(ws, current_row, company_name)
        
        # === RISK SUMMARY ===
        current_row = self._add_risk_summary(ws, current_row, risk_data)
        
        # === RECOMMENDATIONS TABLE ===
        current_row = self._add_recommendations_table(ws, current_row, recommendations)
        
        # === ACTION ITEMS ===
        current_row = self._add_action_items(ws, current_row, recommendations)
        
        # === CHARTS (optional) ===
        if include_charts:
            self._add_priority_chart(ws, recommendations)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    def _add_header(self, ws, start_row: int, company_name: str) -> int:
        """Add header section with branding."""
        # Title
        ws[f'A{start_row}'] = f"{company_name} AI-Generated Recommendations"
        ws[f'A{start_row}'].font = Font(size=18, bold=True, color=self.COLORS['header'])
        ws.merge_cells(f'A{start_row}:G{start_row}')
        
        # Subtitle with date
        start_row += 1
        ws[f'A{start_row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{start_row}'].font = Font(size=10, italic=True, color='666666')
        
        return start_row + 2  # Add spacing
    
    def _add_risk_summary(self, ws, start_row: int, risk_data: Dict) -> int:
        """Add risk assessment summary section."""
        # Section header
        ws[f'A{start_row}'] = "Risk Assessment Summary"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        start_row += 1
        
        # Summary data
        summary_items = [
            ('Risk Score', f"{risk_data.get('risk_score', 'N/A')}/100"),
            ('Risk Level', risk_data.get('risk_level', 'N/A').upper()),
            ('VaR (95%)', f"${risk_data.get('var_95', 0):,.2f}"),
            ('Expected Loss', f"${risk_data.get('expected_loss', 0):,.2f}"),
            ('Confidence', f"{risk_data.get('confidence', 95)}%"),
        ]
        
        for i, (label, value) in enumerate(summary_items):
            row = start_row + i
            ws[f'A{row}'] = label + ":"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            
            # Color code risk level
            if label == 'Risk Level':
                level = risk_data.get('risk_level', '').lower()
                if level in ['high', 'critical']:
                    ws[f'B{row}'].fill = PatternFill(
                        start_color=self.COLORS['high'],
                        fill_type="solid"
                    )
                elif level == 'medium':
                    ws[f'B{row}'].fill = PatternFill(
                        start_color=self.COLORS['medium'],
                        fill_type="solid"
                    )
                else:
                    ws[f'B{row}'].fill = PatternFill(
                        start_color=self.COLORS['low'],
                        fill_type="solid"
                    )
        
        return start_row + len(summary_items) + 2  # Add spacing
    
    def _add_recommendations_table(
        self,
        ws,
        start_row: int,
        recommendations: List[Dict]
    ) -> int:
        """Add recommendations table with formatting."""
        # Section header
        ws[f'A{start_row}'] = "Recommendations"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        start_row += 1
        
        # Table headers
        headers = ['Priority', 'Category', 'Recommendation', 'Impact', 'Effort', 'Est. Cost', 'Timeline']
        
        header_fill = PatternFill(
            start_color=self.COLORS['header'],
            fill_type="solid"
        )
        header_font = Font(bold=True, color=self.COLORS['header_text'])
        center_align = Alignment(horizontal='center', vertical='center')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        start_row += 1
        
        # Table data
        thin_border = Border(
            left=Side(style='thin', color=self.COLORS['border']),
            right=Side(style='thin', color=self.COLORS['border']),
            top=Side(style='thin', color=self.COLORS['border']),
            bottom=Side(style='thin', color=self.COLORS['border'])
        )
        
        alt_fill = PatternFill(
            start_color=self.COLORS['alt_row'],
            fill_type="solid"
        )
        
        for idx, rec in enumerate(recommendations):
            row = start_row + idx
            
            # Data
            ws.cell(row=row, column=1, value=rec.get('priority', 'Medium'))
            ws.cell(row=row, column=2, value=rec.get('category', 'General'))
            ws.cell(row=row, column=3, value=rec.get('text', rec.get('recommendation', '')))
            ws.cell(row=row, column=4, value=rec.get('impact', 'Medium'))
            ws.cell(row=row, column=5, value=rec.get('effort', 'Medium'))
            ws.cell(row=row, column=6, value=rec.get('cost', 'TBD'))
            ws.cell(row=row, column=7, value=rec.get('timeline', 'TBD'))
            
            # Apply formatting
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                
                # Alternating row colors
                if idx % 2 == 1:
                    cell.fill = alt_fill
            
            # Priority color coding
            priority = rec.get('priority', 'Medium')
            priority_cell = ws.cell(row=row, column=1)
            priority_cell.alignment = center_align
            
            if priority.lower() == 'high':
                priority_cell.fill = PatternFill(
                    start_color=self.COLORS['high'],
                    fill_type="solid"
                )
            elif priority.lower() == 'medium':
                priority_cell.fill = PatternFill(
                    start_color=self.COLORS['medium'],
                    fill_type="solid"
                )
            else:
                priority_cell.fill = PatternFill(
                    start_color=self.COLORS['low'],
                    fill_type="solid"
                )
        
        return start_row + len(recommendations) + 2
    
    def _add_action_items(
        self,
        ws,
        start_row: int,
        recommendations: List[Dict]
    ) -> int:
        """Add action items checklist section."""
        # Section header
        ws[f'A{start_row}'] = "Action Items Checklist"
        ws[f'A{start_row}'].font = Font(size=14, bold=True)
        start_row += 1
        
        # High priority items first
        high_priority = [r for r in recommendations if r.get('priority', '').lower() == 'high']
        
        if high_priority:
            ws[f'A{start_row}'] = "⚠️ High Priority (Immediate Action Required)"
            ws[f'A{start_row}'].font = Font(bold=True, color='CC0000')
            start_row += 1
            
            for rec in high_priority:
                ws[f'A{start_row}'] = "☐"
                ws[f'B{start_row}'] = rec.get('text', rec.get('recommendation', ''))[:100]
                start_row += 1
        
        return start_row + 1
    
    def _add_priority_chart(self, ws, recommendations: List[Dict]):
        """Add pie chart showing priority distribution."""
        # Count priorities
        priority_counts = {'High': 0, 'Medium': 0, 'Low': 0}
        for rec in recommendations:
            priority = rec.get('priority', 'Medium')
            if priority in priority_counts:
                priority_counts[priority] += 1
        
        # Create mini data table for chart
        chart_data_start = 50  # Put chart data in column far right
        
        ws.cell(row=1, column=chart_data_start, value='Priority')
        ws.cell(row=1, column=chart_data_start + 1, value='Count')
        
        row = 2
        for priority, count in priority_counts.items():
            ws.cell(row=row, column=chart_data_start, value=priority)
            ws.cell(row=row, column=chart_data_start + 1, value=count)
            row += 1
        
        # Create chart
        chart = PieChart()
        chart.title = "Recommendations by Priority"
        
        data = Reference(ws, min_col=chart_data_start + 1, min_row=1, max_row=4)
        labels = Reference(ws, min_col=chart_data_start, min_row=2, max_row=4)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        
        # Position chart
        ws.add_chart(chart, "I4")


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def extract_recommendations_from_conversation(conversation: List[Dict]) -> List[Dict]:
    """
    Extract structured recommendations from AI conversation.
    
    Parses AI responses to find recommendation patterns.
    """
    recommendations = []
    
    for message in conversation:
        if message.get('role') != 'assistant':
            continue
        
        content = message.get('content', '')
        
        # Look for numbered recommendations
        lines = content.split('\n')
        for line in lines:
            line = line.strip()
            
            # Match patterns like "1. " or "- " or "• "
            if any(line.startswith(prefix) for prefix in ['1.', '2.', '3.', '-', '•', '*']):
                # Extract recommendation text
                text = line.lstrip('0123456789.-•* ')
                
                if len(text) > 10:  # Skip very short lines
                    recommendations.append({
                        'text': text,
                        'priority': _infer_priority(text),
                        'category': _infer_category(text),
                        'impact': 'Medium',
                        'effort': 'Medium',
                        'cost': 'TBD',
                        'timeline': 'TBD'
                    })
    
    return recommendations


def _infer_priority(text: str) -> str:
    """Infer priority from recommendation text."""
    text_lower = text.lower()
    
    high_keywords = ['critical', 'urgent', 'immediately', 'asap', 'high risk', 'must']
    low_keywords = ['consider', 'optional', 'nice to have', 'long-term', 'eventually']
    
    if any(kw in text_lower for kw in high_keywords):
        return 'High'
    elif any(kw in text_lower for kw in low_keywords):
        return 'Low'
    else:
        return 'Medium'


def _infer_category(text: str) -> str:
    """Infer category from recommendation text."""
    text_lower = text.lower()
    
    categories = {
        'Insurance': ['insurance', 'coverage', 'policy', 'claim'],
        'Route': ['route', 'shipping', 'transport', 'carrier'],
        'Packaging': ['packaging', 'container', 'protection'],
        'Timing': ['timing', 'schedule', 'delay', 'date'],
        'Documentation': ['document', 'paperwork', 'customs'],
        'Financial': ['cost', 'price', 'payment', 'budget'],
    }
    
    for category, keywords in categories.items():
        if any(kw in text_lower for kw in keywords):
            return category
    
    return 'General'


# ============================================================
# FASTAPI INTEGRATION EXAMPLE
# ============================================================

"""
Example FastAPI endpoint:

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from app.ai_system_advisor.excel_export import RecommendationExporter

router = APIRouter()

@router.get("/ai/advisor/export/excel/{conversation_id}")
async def export_recommendations_excel(conversation_id: str):
    # Get conversation data
    conversation = get_conversation(conversation_id)
    recommendations = extract_recommendations(conversation)
    risk_data = get_associated_risk_data(conversation_id)
    
    # Generate Excel
    exporter = RecommendationExporter()
    excel_file = exporter.export_recommendations(recommendations, risk_data)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=recommendations_{conversation_id}.xlsx"
        }
    )
"""
